from __future__ import annotations

import csv
import json
import math
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
RAW_ROOT = Path(r"C:\Users\user\Documents\AeronetDATA")
DATA_DIR = ROOT / "aeronet_summary_data"
OUT_PATH = ROOT / "AERONET_SSA_FMF_SDA_statistics_2020-2026.html"
RAW_CSV_PATH = DATA_DIR / "retrieval_individual_observations_no_2022-10-21.csv"
SUMMARY_CSV_PATH = DATA_DIR / "retrieval_monthly_seasonal_statistics.csv"
DESCRIPTIVE_CSV_PATH = DATA_DIR / "retrieval_descriptive_statistics.csv"
EXCLUDED_DATES = {"2022-10-21"}

METRICS = [
    {
        "key": "SSA_440nm",
        "label": "SSA 440 nm",
        "column": "Single_Scattering_Albedo[440nm]",
        "product": "SSA",
        "color": "#287d8e",
    },
    {
        "key": "SSA_675nm",
        "label": "SSA 675 nm",
        "column": "Single_Scattering_Albedo[675nm]",
        "product": "SSA",
        "color": "#3a9278",
    },
    {
        "key": "SSA_870nm",
        "label": "SSA 870 nm",
        "column": "Single_Scattering_Albedo[870nm]",
        "product": "SSA",
        "color": "#5577a5",
    },
    {
        "key": "SSA_1020nm",
        "label": "SSA 1020 nm",
        "column": "Single_Scattering_Albedo[1020nm]",
        "product": "SSA",
        "color": "#7b65a6",
    },
    {
        "key": "FMF_500nm",
        "label": "FMF 500 nm",
        "column": "FineModeFraction_500nm[eta]",
        "product": "SDA",
        "color": "#ba6b2e",
    },
    {
        "key": "AODf_500nm",
        "label": "Fine-mode AOD 500 nm",
        "column": "Fine_Mode_AOD_500nm[tau_f]",
        "product": "SDA",
        "color": "#2f78b7",
    },
    {
        "key": "AODc_500nm",
        "label": "Coarse-mode AOD 500 nm",
        "column": "Coarse_Mode_AOD_500nm[tau_c]",
        "product": "SDA",
        "color": "#b04d52",
    },
]
METRIC_KEYS = [metric["key"] for metric in METRICS]
METRIC_BY_COLUMN = {metric["column"]: metric for metric in METRICS}
SEASONS = {
    1: "Winter",
    2: "Winter",
    3: "Spring",
    4: "Spring",
    5: "Spring",
    6: "Summer",
    7: "Summer",
    8: "Summer",
    9: "Autumn",
    10: "Autumn",
    11: "Autumn",
    12: "Winter",
}
MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def parse_float(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number <= -900:
        return None
    return number


def parse_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d:%m:%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parse_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M:%S")
        except ValueError:
            pass
    return text


def retrieval_files() -> list[tuple[str, Path]]:
    files = []
    for path in RAW_ROOT.rglob("*.xlsx"):
        name = path.name.lower()
        if "no data" in name or "lunar" in name:
            continue
        if name.endswith(".ssa.xlsx"):
            files.append(("SSA", path))
        elif "sda" in name and "aod" not in name:
            files.append(("SDA", path))
    return sorted(files, key=lambda item: str(item[1]).lower())


def find_header(row: tuple, product: str) -> dict[str, int] | None:
    headers = [str(value).strip() if value is not None else "" for value in row]
    date_index = next((i for i, value in enumerate(headers) if value.lower().startswith("date")), None)
    if date_index is None:
        return None
    indexes = {"date": date_index}
    time_index = next((i for i, value in enumerate(headers) if value.lower().startswith("time")), None)
    if time_index is not None:
        indexes["time"] = time_index
    for column, metric in METRIC_BY_COLUMN.items():
        if metric["product"] == product and column in headers:
            indexes[metric["key"]] = headers.index(column)
    return indexes if len(indexes) > (2 if "time" in indexes else 1) else None


def scan_records():
    records_by_key: dict[tuple, dict] = {}
    scanned_files = []
    skipped_files = []
    product_counts = defaultdict(int)

    for product, path in retrieval_files():
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            skipped_files.append({"file": str(path), "reason": str(exc)})
            continue
        found = 0
        for sheet in workbook.worksheets:
            if str(sheet.title).strip() == "-999":
                continue
            header = None
            for row in sheet.iter_rows(values_only=True):
                if header is None:
                    header = find_header(row, product)
                    continue
                date = parse_date(row[header["date"]] if header["date"] < len(row) else None)
                if not date or date in EXCLUDED_DATES:
                    continue
                time_index = header.get("time")
                time = parse_time(row[time_index] if time_index is not None and time_index < len(row) else None)
                values = {}
                for metric in METRICS:
                    index = header.get(metric["key"])
                    if index is None or index >= len(row):
                        continue
                    number = parse_float(row[index])
                    if number is not None:
                        values[metric["key"]] = number
                if not values:
                    continue
                key = (product, date, time)
                record = records_by_key.setdefault(
                    key,
                    {"date": date, "time": time, "product": product, **{metric: None for metric in METRIC_KEYS}},
                )
                for metric, number in values.items():
                    if record[metric] is None:
                        record[metric] = number
                found += 1
        workbook.close()
        if found:
            scanned_files.append(str(path))
            product_counts[product] += found
        else:
            skipped_files.append({"file": str(path), "reason": "No usable retrieval observations"})

    records = sorted(records_by_key.values(), key=lambda row: (row["date"], row["time"], row["product"]))
    return records, scanned_files, skipped_files, dict(product_counts)


def daily_records(records):
    grouped = defaultdict(lambda: defaultdict(list))
    for record in records:
        for metric in METRIC_KEYS:
            value = record.get(metric)
            if value is not None:
                grouped[record["date"]][metric].append(value)
    output = []
    for date in sorted(grouped):
        row = {"date": date}
        for metric in METRIC_KEYS:
            values = grouped[date].get(metric, [])
            row[metric] = statistics.fmean(values) if values else None
        output.append(row)
    return output


def descriptive(values):
    clean = sorted(value for value in values if value is not None)
    if not clean:
        return {"n": 0, "mean": None, "std": None, "median": None, "min": None, "max": None, "skewness": None}
    count = len(clean)
    mean = statistics.fmean(clean)
    std = statistics.stdev(clean) if count > 1 else 0.0
    skewness = None
    if count >= 3:
        skewness = 0.0 if std == 0 else (
            count / ((count - 1) * (count - 2))
            * sum(((value - mean) / std) ** 3 for value in clean)
        )
    return {
        "n": count,
        "mean": mean,
        "std": std,
        "median": statistics.median(clean),
        "min": clean[0],
        "max": clean[-1],
        "skewness": skewness,
    }


def compact_rows(records):
    return [
        [record["date"], *[round(record[metric], 6) if record.get(metric) is not None else None for metric in METRIC_KEYS]]
        for record in records
    ]


def group_value(date_text: str, grouping: str) -> str:
    date = datetime.strptime(date_text, "%Y-%m-%d")
    if grouping == "month":
        return MONTH_NAMES[date.month - 1]
    if grouping == "season":
        return SEASONS[date.month]
    if grouping == "year":
        return str(date.year)
    return f"{date.year} {SEASONS[date.month]}"


def write_exports(records, daily):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with RAW_CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        fields = ["date", "time", "product", *METRIC_KEYS]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)

    descriptive_rows = []
    for basis, basis_records in (("individual", records), ("daily_means", daily)):
        for metric in METRICS:
            stats = descriptive([row.get(metric["key"]) for row in basis_records])
            descriptive_rows.append({"basis": basis, "metric": metric["key"], "label": metric["label"], **stats})
    with DESCRIPTIVE_CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(descriptive_rows[0].keys()))
        writer.writeheader()
        writer.writerows(descriptive_rows)

    summary_rows = []
    for basis, basis_records in (("individual", records), ("daily_means", daily)):
        for grouping in ("month", "season", "year", "year_season"):
            for metric in METRICS:
                grouped = defaultdict(list)
                for record in basis_records:
                    value = record.get(metric["key"])
                    if value is not None:
                        grouped[group_value(record["date"], grouping)].append(value)
                for group, values in grouped.items():
                    stats = descriptive(values)
                    summary_rows.append(
                        {
                            "basis": basis,
                            "grouping": grouping,
                            "group": group,
                            "metric": metric["key"],
                            "label": metric["label"],
                            "n": stats["n"],
                            "mean": stats["mean"],
                            "standard_deviation": stats["std"],
                            "median": stats["median"],
                            "minimum": stats["min"],
                            "maximum": stats["max"],
                            "skewness": stats["skewness"],
                        }
                    )
    with SUMMARY_CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(summary_rows[0].keys()))
        writer.writeheader()
        writer.writerows(summary_rows)


def build_html(payload):
    data_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    template = r'''<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>AERONET SSA, FMF and SDA Statistics</title>
<style>
:root { --ink:#172234; --muted:#566579; --line:#d9e0ea; --paper:#f5f7fa; --panel:#fff; --accent:#26364e; }
* { box-sizing:border-box; }
body { margin:0; background:var(--paper); color:var(--ink); font-family:"Segoe UI",Arial,sans-serif; }
header { padding:20px 24px 14px; background:#fff; border-bottom:1px solid var(--line); }
h1 { margin:0 0 5px; font-size:28px; letter-spacing:0; }
h2,h3 { letter-spacing:0; }
p { margin:0; color:var(--muted); line-height:1.45; }
main { max-width:1180px; padding:16px 24px 28px; display:grid; gap:12px; }
section { min-width:0; background:var(--panel); border:1px solid var(--line); border-radius:8px; overflow:hidden; }
.section-head { padding:13px 14px; border-bottom:1px solid var(--line); display:flex; justify-content:space-between; gap:12px; align-items:start; }
.section-head h2 { margin:0 0 3px; font-size:18px; }
.controls { padding:12px 14px; display:grid; gap:12px; }
.control-row { display:flex; flex-wrap:wrap; gap:8px 14px; align-items:center; }
label { color:var(--muted); font-size:13px; font-weight:700; }
select,button { min-height:34px; border:1px solid var(--line); border-radius:6px; background:#fff; color:var(--ink); font:inherit; }
select { padding:5px 32px 5px 9px; }
button { padding:5px 10px; cursor:pointer; font-weight:700; }
button.active { background:var(--accent); border-color:var(--accent); color:#fff; }
.period-wrap { overflow-x:auto; padding-bottom:4px; }
.period-grid { min-width:840px; display:grid; grid-template-columns:82px repeat(12,minmax(48px,1fr)); gap:6px; align-items:center; }
.period-grid span { text-align:center; color:var(--muted); font-size:12px; font-weight:700; }
.period-grid button:disabled { background:#f1f3f6; color:#a4adba; cursor:default; }
.status { color:var(--muted); font-size:12px; }
.chart-scroll { width:100%; overflow-x:auto; padding:12px 14px 4px; }
.chart-wrap { position:relative; min-width:100%; height:430px; }
canvas { display:block; height:420px; }
.tooltip { position:absolute; display:none; z-index:4; width:230px; padding:8px 9px; border:1px solid #b9c5d5; border-radius:6px; background:rgba(255,255,255,.97); box-shadow:0 8px 24px rgba(17,24,39,.16); font-size:12px; line-height:1.4; pointer-events:none; }
.tooltip strong { display:block; margin-bottom:3px; }
.table-wrap { max-height:430px; overflow:auto; }
table { width:100%; border-collapse:collapse; font-size:12px; }
th,td { padding:8px 10px; border-bottom:1px solid #edf1f6; text-align:right; white-space:nowrap; }
th:first-child,td:first-child { text-align:left; }
th { position:sticky; top:0; z-index:1; background:#f9fbfd; color:var(--muted); }
.note { padding:10px 14px 13px; font-size:12px; }
@media (max-width:720px) { header,main { padding-left:12px; padding-right:12px; } .section-head { display:grid; } }
</style>
</head>
<body>
<header>
  <h1>SSA, FMF and fine/coarse AOD statistics</h1>
  <p>Interactive individual-measurement and daily-mean statistics. Solar SDA and inversion data only; LUNAR, NO DATA and 2022-10-21 are excluded.</p>
</header>
<main>
  <section>
    <div class="section-head"><div><h2>Data selector</h2><p>Choose exact calendar months. The selection is applied to every chart and table below.</p></div><div class="status" id="selectionStatus"></div></div>
    <div class="controls">
      <div class="control-row"><button id="keepAll">Keep all</button><button id="removeAll">Remove all</button></div>
      <div class="period-wrap"><div class="period-grid" id="periodGrid"></div></div>
    </div>
  </section>
  <section id="retrieval-distributions">
    <div class="section-head"><div><h2 id="chartTitle">Monthly distribution</h2><p id="chartMeta"></p></div></div>
    <div class="controls">
      <div class="control-row"><label for="metricSelect">Metric</label><select id="metricSelect"></select></div>
      <div class="control-row"><label>Basis</label><button id="individualBtn" class="active">Individual measurements</button><button id="dailyBtn">Daily means</button></div>
      <div class="control-row"><label>Grouping</label><button data-group="month" class="active">Calendar month</button><button data-group="season">Season</button><button data-group="year">Year</button><button data-group="year_season">Season of each year</button></div>
    </div>
    <div class="chart-scroll" id="chartScroll"><div class="chart-wrap" id="chartWrap"><canvas id="chart"></canvas><div class="tooltip" id="tooltip"></div></div></div>
    <p class="note">Bars show means and whiskers show mean plus/minus sample standard deviation. Hover over a bar for its values.</p>
  </section>
  <section id="retrieval-statistics">
    <div class="section-head"><div><h2>Descriptive statistics</h2><p id="statsMeta"></p></div></div>
    <div class="table-wrap"><table><thead><tr><th>Metric</th><th>n</th><th>Mean</th><th>Standard deviation</th><th>Median</th><th>Minimum</th><th>Maximum</th><th>Skewness</th></tr></thead><tbody id="statsRows"></tbody></table></div>
    <p class="note">Skewness is adjusted Fisher-Pearson sample skewness. Each metric is calculated separately from its own available values.</p>
  </section>
  <section>
    <div class="section-head"><div><h2>Grouped values for the selected metric</h2><p id="tableMeta"></p></div></div>
    <div class="table-wrap"><table><thead><tr><th>Group</th><th>n</th><th>Mean</th><th>Standard deviation</th><th>Median</th><th>Minimum</th><th>Maximum</th><th>Skewness</th></tr></thead><tbody id="groupRows"></tbody></table></div>
  </section>
</main>
<script>
const DATA=__DATA__;
const metricIndex=key=>DATA.metric_keys.indexOf(key)+1;
const periods=[...new Set(DATA.individual.map(row=>row[0].slice(0,7)))].sort();
const periodSet=new Set(periods);
const years=[...new Set(periods.map(value=>Number(value.slice(0,4))))].sort((a,b)=>a-b);
const months=[[1,"Jan"],[2,"Feb"],[3,"Mar"],[4,"Apr"],[5,"May"],[6,"Jun"],[7,"Jul"],[8,"Aug"],[9,"Sep"],[10,"Oct"],[11,"Nov"],[12,"Dec"]];
const state={metric:DATA.metric_keys[0],basis:"individual",group:"month",selected:new Set(initialPeriods())};
let layout={bars:[]};

function initialPeriods(){
  const query=new URLSearchParams(location.search).get("periods");
  if(!query)return periods;
  const selected=query.split(",").map(x=>x.trim()).filter(x=>periodSet.has(x));
  return selected;
}
function metricInfo(key){return DATA.metrics.find(metric=>metric.key===key);}
function filteredRows(){const rows=state.basis==="individual"?DATA.individual:DATA.daily;return rows.filter(row=>state.selected.has(row[0].slice(0,7)));}
function valuesFor(metric){const idx=metricIndex(metric);return filteredRows().map(row=>row[idx]).filter(Number.isFinite);}
function describe(values){
  const clean=values.filter(Number.isFinite).sort((a,b)=>a-b),n=clean.length;
  if(!n)return {n:0,mean:null,std:null,median:null,min:null,max:null,skewness:null};
  const mean=clean.reduce((a,b)=>a+b,0)/n;
  const variance=n>1?clean.reduce((sum,v)=>sum+(v-mean)**2,0)/(n-1):0;
  const std=Math.sqrt(variance),middle=Math.floor(n/2),median=n%2?clean[middle]:(clean[middle-1]+clean[middle])/2;
  const skewness=n<3?null:(std===0?0:n/((n-1)*(n-2))*clean.reduce((sum,v)=>sum+((v-mean)/std)**3,0));
  return {n,mean,std,median,min:clean[0],max:clean[n-1],skewness};
}
function season(month){return month<=2||month===12?"Winter":month<=5?"Spring":month<=8?"Summer":"Autumn";}
function groupKey(date){const year=Number(date.slice(0,4)),month=Number(date.slice(5,7));if(state.group==="month")return DATA.month_names[month-1];if(state.group==="season")return season(month);if(state.group==="year")return String(year);return `${year} ${season(month)}`;}
function groupOrder(key){
  if(state.group==="month")return DATA.month_names.indexOf(key);
  if(state.group==="season")return ["Winter","Spring","Summer","Autumn"].indexOf(key);
  if(state.group==="year")return Number(key);
  const [year,label]=key.split(" ");return Number(year)*10+["Winter","Spring","Summer","Autumn"].indexOf(label);
}
function grouped(){
  const idx=metricIndex(state.metric),map=new Map();
  filteredRows().forEach(row=>{const value=row[idx];if(!Number.isFinite(value))return;const key=groupKey(row[0]);if(!map.has(key))map.set(key,[]);map.get(key).push(value);});
  return [...map].sort((a,b)=>groupOrder(a[0])-groupOrder(b[0])).map(([group,values])=>({group,...describe(values)}));
}
function fmt(value,digits=5){return value===null||value===undefined||!Number.isFinite(value)?"n/a":Number(value).toFixed(digits);}
function renderPeriods(){
  const cells=[document.createElement("span")];months.forEach(([,label])=>{const span=document.createElement("span");span.textContent=label;cells.push(span);});
  years.forEach(year=>{
    const yearButton=document.createElement("button");yearButton.textContent=String(year);const yearPeriods=periods.filter(p=>p.startsWith(`${year}-`));yearButton.classList.toggle("active",yearPeriods.every(p=>state.selected.has(p)));yearButton.onclick=()=>{const keep=!yearPeriods.every(p=>state.selected.has(p));yearPeriods.forEach(p=>keep?state.selected.add(p):state.selected.delete(p));renderAll();};cells.push(yearButton);
    months.forEach(([month,label])=>{const period=`${year}-${String(month).padStart(2,"0")}`,button=document.createElement("button");button.textContent=label;button.disabled=!periodSet.has(period);button.classList.toggle("active",state.selected.has(period));button.onclick=()=>{state.selected.has(period)?state.selected.delete(period):state.selected.add(period);renderAll();};cells.push(button);});
  });document.getElementById("periodGrid").replaceChildren(...cells);
}
function renderStats(){
  const rows=DATA.metric_keys.map(key=>{const stats=describe(valuesFor(key)),tr=document.createElement("tr");tr.innerHTML=`<td>${metricInfo(key).label}</td><td>${stats.n.toLocaleString()}</td><td>${fmt(stats.mean)}</td><td>${fmt(stats.std)}</td><td>${fmt(stats.median)}</td><td>${fmt(stats.min)}</td><td>${fmt(stats.max)}</td><td>${fmt(stats.skewness)}</td>`;return tr;});
  document.getElementById("statsRows").replaceChildren(...rows);document.getElementById("statsMeta").textContent=`${state.basis==="individual"?"Individual measurements":"Daily means"}; ${state.selected.size} calendar months selected.`;
}
function renderTable(rows){document.getElementById("groupRows").innerHTML=rows.map(row=>`<tr><td>${row.group}</td><td>${row.n.toLocaleString()}</td><td>${fmt(row.mean)}</td><td>${fmt(row.std)}</td><td>${fmt(row.median)}</td><td>${fmt(row.min)}</td><td>${fmt(row.max)}</td><td>${fmt(row.skewness)}</td></tr>`).join("");document.getElementById("tableMeta").textContent=`${metricInfo(state.metric).label}; ${state.basis==="individual"?"individual measurements":"daily means"}.`;}
function draw(rows){
  const canvas=document.getElementById("chart"),wrap=document.getElementById("chartWrap"),dpr=devicePixelRatio||1,width=Math.max(document.getElementById("chartScroll").clientWidth-28,rows.length*58+90,680),height=420;wrap.style.width=`${width}px`;canvas.style.width=`${width}px`;canvas.style.height=`${height}px`;canvas.width=Math.floor(width*dpr);canvas.height=Math.floor(height*dpr);const ctx=canvas.getContext("2d");ctx.setTransform(dpr,0,0,dpr,0,0);ctx.clearRect(0,0,width,height);
  const margin={left:64,right:18,top:24,bottom:84},plotW=width-margin.left-margin.right,plotH=height-margin.top-margin.bottom,maxY=Math.max(...rows.map(r=>(r.mean||0)+(r.std||0)),1e-6)*1.12,y=v=>margin.top+plotH-v/maxY*plotH,band=plotW/Math.max(rows.length,1),barW=Math.min(38,band*.62),info=metricInfo(state.metric);layout={bars:[]};ctx.font="12px Segoe UI, Arial";
  ctx.strokeStyle="#dde3eb";ctx.fillStyle="#566579";ctx.textAlign="right";ctx.textBaseline="middle";for(let i=0;i<=5;i++){const value=maxY*i/5,yy=y(value);ctx.beginPath();ctx.moveTo(margin.left,yy);ctx.lineTo(width-margin.right,yy);ctx.stroke();ctx.fillText(value.toFixed(info.key.startsWith("AOD")?2:3),margin.left-8,yy);}
  if(!rows.length){ctx.textAlign="left";ctx.fillText("No data for this selection",margin.left,margin.top+24);return;}
  rows.forEach((row,i)=>{const x=margin.left+i*band+(band-barW)/2,top=y(row.mean),base=y(0),center=x+barW/2,errTop=y(row.mean+row.std),errBottom=y(Math.max(0,row.mean-row.std));ctx.fillStyle=info.color;ctx.fillRect(x,top,barW,base-top);ctx.strokeStyle="#263241";ctx.beginPath();ctx.moveTo(center,errTop);ctx.lineTo(center,errBottom);ctx.moveTo(center-6,errTop);ctx.lineTo(center+6,errTop);ctx.moveTo(center-6,errBottom);ctx.lineTo(center+6,errBottom);ctx.stroke();ctx.save();ctx.translate(center,base+12);ctx.rotate(-Math.PI/4);ctx.fillStyle="#566579";ctx.textAlign="right";ctx.textBaseline="middle";ctx.fillText(row.group,0,0);ctx.restore();layout.bars.push({x,y:errTop,w:barW,h:base-errTop,row});});
  ctx.strokeStyle="#687587";ctx.beginPath();ctx.moveTo(margin.left,margin.top);ctx.lineTo(margin.left,margin.top+plotH);ctx.lineTo(width-margin.right,margin.top+plotH);ctx.stroke();
}
function renderAll(){
  renderPeriods();const rows=grouped();renderStats();renderTable(rows);draw(rows);const groupLabel={month:"calendar month",season:"season",year:"year",year_season:"season of each year"}[state.group];document.getElementById("chartTitle").textContent=`${metricInfo(state.metric).label} by ${groupLabel}`;document.getElementById("chartMeta").textContent=`${state.basis==="individual"?"Individual-measurement":"Daily-mean"} basis; mean plus/minus standard deviation.`;document.getElementById("selectionStatus").textContent=`${state.selected.size} of ${periods.length} available calendar months kept.`;document.getElementById("individualBtn").classList.toggle("active",state.basis==="individual");document.getElementById("dailyBtn").classList.toggle("active",state.basis==="daily");document.querySelectorAll("[data-group]").forEach(btn=>btn.classList.toggle("active",btn.dataset.group===state.group));document.getElementById("tooltip").style.display="none";
}
DATA.metrics.forEach(metric=>{const option=document.createElement("option");option.value=metric.key;option.textContent=metric.label;document.getElementById("metricSelect").append(option);});
document.getElementById("metricSelect").onchange=event=>{state.metric=event.target.value;renderAll();};
document.getElementById("individualBtn").onclick=()=>{state.basis="individual";renderAll();};document.getElementById("dailyBtn").onclick=()=>{state.basis="daily";renderAll();};document.querySelectorAll("[data-group]").forEach(btn=>btn.onclick=()=>{state.group=btn.dataset.group;renderAll();});document.getElementById("keepAll").onclick=()=>{state.selected=new Set(periods);renderAll();};document.getElementById("removeAll").onclick=()=>{state.selected.clear();renderAll();};
document.getElementById("chart").addEventListener("mousemove",event=>{const rect=event.currentTarget.getBoundingClientRect(),x=event.clientX-rect.left,y=event.clientY-rect.top,bar=layout.bars.find(item=>x>=item.x-5&&x<=item.x+item.w+5&&y>=item.y-8&&y<=item.y+item.h+8),tip=document.getElementById("tooltip");if(!bar){tip.style.display="none";return;}tip.innerHTML=`<strong>${bar.row.group}</strong><div>Mean: ${fmt(bar.row.mean)}</div><div>Standard deviation: ${fmt(bar.row.std)}</div><div>n: ${bar.row.n.toLocaleString()}</div>`;tip.style.left=`${Math.min(Math.max(8,x+12),rect.width-238)}px`;tip.style.top=`${Math.max(8,y-72)}px`;tip.style.display="block";});document.getElementById("chart").addEventListener("mouseleave",()=>document.getElementById("tooltip").style.display="none");window.addEventListener("resize",()=>draw(grouped()));renderAll();
</script>
</body>
</html>'''
    OUT_PATH.write_text(template.replace("__DATA__", data_json), encoding="utf-8")


def main():
    records, source_files, skipped_files, product_counts = scan_records()
    daily = daily_records(records)
    write_exports(records, daily)
    dates = sorted({record["date"] for record in records})
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_dir": str(RAW_ROOT),
        "source_files": len(source_files),
        "skipped_files": len(skipped_files),
        "product_counts": product_counts,
        "excluded_dates": sorted(EXCLUDED_DATES),
        "date_range": [dates[0], dates[-1]] if dates else ["", ""],
        "metric_keys": METRIC_KEYS,
        "metrics": [{key: metric[key] for key in ("key", "label", "product", "color")} for metric in METRICS],
        "month_names": MONTH_NAMES,
        "individual": compact_rows(records),
        "daily": compact_rows(daily),
    }
    build_html(payload)
    print(f"Wrote {OUT_PATH}")
    print(f"Source files: {len(source_files)}; skipped: {len(skipped_files)}")
    print(f"Individual rows: {len(records)}; daily rows: {len(daily)}")
    print(f"Product rows scanned: {product_counts}")


if __name__ == "__main__":
    main()
