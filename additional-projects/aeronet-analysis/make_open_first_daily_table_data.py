from __future__ import annotations

import csv
import json
import math
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
WORKSPACE = BASE.parent
RAW_ROOT = Path(r"C:\Users\user\Documents\AeronetDATA")
DAILY_SUMMARY = BASE / "aeronet_summary_data_no_2022-10-21" / "daily_summary.csv"
OUTPUT_JS = BASE / "open_first_daily_table_data.js"
EXCLUDED_DATES = {"2022-10-21"}

FIELD_COLUMNS = {
    "aodf500": "Fine_Mode_AOD_500nm[tau_f]",
    "aodc500": "Coarse_Mode_AOD_500nm[tau_c]",
    "sf": "Sphericity_Factor(%)",
    "dr440": "Depolarization_Ratio[440nm]",
    "nr440": "Refractive_Index-Real_Part[440nm]",
}


def parse_float(value) -> float | None:
    if value is None:
        return None
    try:
        number = float(str(value).strip())
    except ValueError:
        return None
    if not math.isfinite(number) or number <= -900:
        return None
    return number


def parse_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d:%m:%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def find_header(row: tuple) -> dict[str, int] | None:
    headers = [str(value).strip() if value is not None else "" for value in row]
    date_col = next((i for i, name in enumerate(headers) if name.lower().startswith("date")), None)
    if date_col is None:
        return None
    indexes = {"date": date_col}
    for field, column_name in FIELD_COLUMNS.items():
        if column_name in headers:
            indexes[field] = headers.index(column_name)
    return indexes if len(indexes) > 1 else None


def scan_retrieval_workbooks() -> dict[str, dict[str, list[float]]]:
    values: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    if not RAW_ROOT.exists():
        return values

    for workbook_path in RAW_ROOT.rglob("*.xlsx"):
        name = workbook_path.name.upper()
        suffix = workbook_path.name.lower()
        has_retrieval_fields = (
            "SDA" in name
            or suffix.endswith(".asy.xlsx")
            or suffix.endswith(".lid.xlsx")
            or suffix.endswith(".rin.xlsx")
        )
        if "NO DATA" in name or not has_retrieval_fields:
            continue
        try:
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        except Exception:
            continue
        for sheet in workbook.worksheets:
            if str(sheet.title).strip() == "-999":
                continue
            header = None
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if header is None:
                    header = find_header(row)
                    continue
                date = parse_date(row[header["date"]] if header["date"] < len(row) else None)
                if not date or date in EXCLUDED_DATES:
                    continue
                for field in FIELD_COLUMNS:
                    column_index = header.get(field)
                    if column_index is None or column_index >= len(row):
                        continue
                    number = parse_float(row[column_index])
                    if number is not None:
                        values[date][field].append(number)
    return values


def mean_std(values: list[float]) -> tuple[float, float]:
    mean = statistics.fmean(values)
    std = statistics.stdev(values) if len(values) > 1 else 0.0
    return mean, std


def fmt_number(value: float, decimals: int) -> str:
    return f"{value:.{decimals}f}"


def fmt_mean_std(values: list[float]) -> str | None:
    if not values:
        return None
    mean, std = mean_std(values)
    std_decimals = 3 if 0 < abs(std) < 0.01 else 2
    return f"{fmt_number(mean, 2)} &plusmn; {fmt_number(std, std_decimals)}"


def fmt_range(values: list[float], decimals: int) -> str | None:
    if not values:
        return None
    low = min(values)
    high = max(values)
    if round(low, decimals) == round(high, decimals):
        return fmt_number(low, decimals)
    return f"{fmt_number(low, decimals)}&ndash;{fmt_number(high, decimals)}"


def fmt_daily_date(iso_date: str) -> str:
    date = datetime.strptime(iso_date, "%Y-%m-%d")
    return f"{date.day} {date.strftime('%B %Y')}"


def fmt_summary_mean_std(row: dict[str, str], prefix: str) -> str:
    mean = parse_float(row.get(f"{prefix}_mean"))
    std = parse_float(row.get(f"{prefix}_std"))
    if mean is None:
        return "Unavailable"
    if std is None:
        std = 0.0
    return f"{fmt_number(mean, 2)} &plusmn; {fmt_number(std, 2)}"


def build_rows() -> list[dict[str, object]]:
    retrieval = scan_retrieval_workbooks()
    rows = []
    with DAILY_SUMMARY.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            iso_date = row["date"]
            if iso_date in EXCLUDED_DATES:
                continue
            extra = retrieval.get(iso_date, {})
            rows.append(
                {
                    "date": fmt_daily_date(iso_date),
                    "iso": iso_date,
                    "month": int(row["month"]),
                    "aod500Mean": parse_float(row.get("AOD_500nm_mean")) or 0,
                    "aod440": fmt_summary_mean_std(row, "AOD_440nm"),
                    "aod500": fmt_summary_mean_std(row, "AOD_500nm"),
                    "aodf500": fmt_mean_std(extra.get("aodf500", [])),
                    "aodc500": fmt_mean_std(extra.get("aodc500", [])),
                    "ae440870": fmt_summary_mean_std(row, "440-870_Angstrom_Exponent"),
                    "ae380500": fmt_summary_mean_std(row, "380-500_Angstrom_Exponent"),
                    "sf": fmt_range(extra.get("sf", []), 1),
                    "dr440": fmt_range(extra.get("dr440", []), 3),
                    "nr440": fmt_range(extra.get("nr440", []), 2),
                }
            )
    return rows


def main() -> None:
    rows = build_rows()
    OUTPUT_JS.write_text(
        "window.AERONET_DAILY_TABLE = "
        + json.dumps(rows, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    filled = {
        field: sum(1 for row in rows if row.get(field))
        for field in ("aodf500", "aodc500", "sf", "dr440", "nr440")
    }
    print(f"Wrote {OUTPUT_JS}")
    print(f"Rows: {len(rows)}")
    print(f"Filled retrieval fields: {filled}")


if __name__ == "__main__":
    main()
